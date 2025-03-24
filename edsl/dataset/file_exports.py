from abc import ABC, abstractmethod
import io
import csv
import base64
import sqlite3
from typing import Optional, Union, Any, Dict


class FileExport(ABC):
    def __init__(
        self,
        data: Any,
        filename: Optional[str] = None,
        remove_prefix: bool = False,
        pretty_labels: Optional[Dict[str, str]] = None,
    ):
        self.data = data
        self.filename = filename  # or self._get_default_filename()
        self.remove_prefix = remove_prefix
        self.pretty_labels = pretty_labels

    @property
    def mime_type(self) -> str:
        """Return the MIME type for this export format."""
        return self.__class__.mime_type

    @property
    def suffix(self) -> str:
        """Return the file suffix for this format."""
        return self.__class__.suffix

    @property
    def is_binary(self) -> bool:
        """Whether the format is binary or text-based."""
        return self.__class__.is_binary

    def _get_default_filename(self) -> str:
        """Generate default filename for this format."""
        return f"results.{self.suffix}"

    def _create_filestore(self, data: Union[str, bytes]):
        """Create a FileStore instance with encoded data."""
        from ..scenarios.file_store import FileStore
        if isinstance(data, str):
            base64_string = base64.b64encode(data.encode()).decode()
        else:
            base64_string = base64.b64encode(data).decode()

        # FileStore already imported

        path = self.filename or self._get_default_filename()

        fs = FileStore(
            path=path,
            mime_type=self.mime_type,
            binary=self.is_binary,
            suffix=self.suffix,
            base64_string=base64_string,
        )

        if self.filename is not None:
            fs.write(self.filename)
            return None
        return fs

    @abstractmethod
    def format_data(self) -> Union[str, bytes]:
        """Convert the input data to the target format."""
        pass

    def export(self) -> Optional:
        """Export the data to a FileStore instance.
        
        Returns:
            A FileStore instance or None if the file was written directly.
        """
        formatted_data = self.format_data()
        return self._create_filestore(formatted_data)


class JSONLExport(FileExport):
    mime_type = "application/jsonl"
    suffix = "jsonl"
    is_binary = False

    def format_data(self) -> str:
        output = io.StringIO()
        for entry in self.data:
            key, values = list(entry.items())[0]
            output.write(f'{{"{key}": {values}}}\n')
        return output.getvalue()


class TabularExport(FileExport, ABC):
    """Base class for exports that use tabular data."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.header, self.rows = self.data.get_tabular_data(
            remove_prefix=self.remove_prefix, pretty_labels=self.pretty_labels
        )


class CSVExport(TabularExport):
    mime_type = "text/csv"
    suffix = "csv"
    is_binary = False

    def format_data(self) -> str:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(self.header)
        writer.writerows(self.rows)
        return output.getvalue()


class ExcelExport(TabularExport):
    mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    suffix = "xlsx"
    is_binary = True

    def __init__(self, *args, sheet_name: Optional[str] = None, **kwargs):
        super().__init__(*args, **kwargs)
        self.sheet_name = sheet_name or "Results"

    def format_data(self) -> bytes:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = self.sheet_name

        # Write header
        for col, value in enumerate(self.header, 1):
            ws.cell(row=1, column=col, value=value)

        # Write data rows
        for row_idx, row_data in enumerate(self.rows, 2):
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col, value=value)

        # Save to bytes buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()




class SQLiteExport(TabularExport):
    mime_type = "application/x-sqlite3"
    suffix = "db"
    is_binary = True

    def __init__(
        self, *args, table_name: str = "results", if_exists: str = "replace", **kwargs
    ):
        """
        Initialize SQLite export.

        Args:
            table_name: Name of the table to create
            if_exists: How to handle existing table ('fail', 'replace', or 'append')
        """
        super().__init__(*args, **kwargs)
        self.table_name = table_name
        self.if_exists = if_exists

    def _get_column_types(self) -> list[tuple[str, str]]:
        """Infer SQL column types from the data."""
        column_types = []

        # Check first row of data for types
        if self.rows:
            first_row = self.rows[0]
            for header, value in zip(self.header, first_row):
                if isinstance(value, bool):
                    sql_type = "BOOLEAN"
                elif isinstance(value, int):
                    sql_type = "INTEGER"
                elif isinstance(value, float):
                    sql_type = "REAL"
                else:
                    sql_type = "TEXT"
                column_types.append((header, sql_type))
        else:
            # If no data, default to TEXT
            column_types = [(header, "TEXT") for header in self.header]

        return column_types

    def _create_table(self, cursor: sqlite3.Cursor) -> None:
        """Create the table with appropriate schema."""
        column_types = self._get_column_types()

        # Drop existing table if replace mode
        if self.if_exists == "replace":
            cursor.execute(f"DROP TABLE IF EXISTS {self.table_name}")
        elif self.if_exists == "fail":
            cursor.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                (self.table_name,),
            )
            if cursor.fetchone():
                from .exceptions import DatasetValueError
                raise DatasetValueError(f"Table {self.table_name} already exists")

        # Create table
        columns = ", ".join(f'"{col}" {dtype}' for col, dtype in column_types)
        create_table_sql = f"""
        CREATE TABLE IF NOT EXISTS {self.table_name} (
            {columns}
        )
        """
        cursor.execute(create_table_sql)

    def format_data(self) -> bytes:
        """Convert the data to a SQLite database file."""
        buffer = io.BytesIO()

        # Create in-memory database
        conn = sqlite3.connect(":memory:")
        cursor = conn.cursor()

        # Create table and insert data
        self._create_table(cursor)

        # Prepare placeholders for INSERT
        placeholders = ",".join(["?" for _ in self.header])
        insert_sql = f"INSERT INTO {self.table_name} ({','.join(self.header)}) VALUES ({placeholders})"

        # Insert data
        cursor.executemany(insert_sql, self.rows)
        conn.commit()

        # Save to file buffer
        conn.backup(sqlite3.connect(buffer))
        conn.close()

        buffer.seek(0)
        return buffer.getvalue()

    def _validate_params(self) -> None:
        """Validate initialization parameters."""
        valid_if_exists = {"fail", "replace", "append"}
        if self.if_exists not in valid_if_exists:
            from .exceptions import DatasetValueError
            raise DatasetValueError(
                f"if_exists must be one of {valid_if_exists}, got {self.if_exists}"
            )

        # Validate table name (basic SQLite identifier validation)
        if not self.table_name.isalnum() and not all(c in "_" for c in self.table_name):
            from .exceptions import DatasetValueError
            raise DatasetValueError(
                f"Invalid table name: {self.table_name}. Must contain only alphanumeric characters and underscores."
            )
